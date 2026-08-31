#!/usr/bin/env python3
"""Apply Internal Audit corrections to the ECAG Transaction Fees & Incentives workbook.

Reads the '_To_be_corrected' workbook, applies the change spec derived from the
audit-findings workbook (IA comments col J, ANP responses col K, late notes col N,
audit row N == target row N-1), and writes a corrected workbook plus an audit report.

Every edit asserts the current cell value first; the script fails loudly on mismatch.
Row deletions are done bottom-up with manual re-application of merged ranges and
row heights (openpyxl does not shift those reliably).
"""
import copy
import datetime
import sys
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Font, Alignment

SRC = "/Users/nishantkanchan/Downloads/WS - Applied/Deutsche Borse/Office-Work/01_Input/Overview of Transaction Fees and Incentives (ECAG)_To_be_corrected.xlsx"
OUTDIR = "/Users/nishantkanchan/Downloads/WS - Applied/Deutsche Borse/Office-Work/02_Output"
TODAY = "2026-09-01"
OUT_WB = f"{OUTDIR}/Overview of Transaction Fees and Incentives (ECAG)_Corrected_{TODAY}.xlsx"
OUT_REPORT = f"{OUTDIR}/Audit_Correction_Report_{TODAY}.xlsx"

# ---------------------------------------------------------------------------
# Change specification. target_row = audit_row - 1 (audit file has extra title row).
# Each edit: (cell, expected_old, new, audit_row, ia_comment, anp_response, reason, review_note)
EDITS = [
    ("C22",
     "Mini-Futures on the MDAX® (FSMX)",
     "Futures on Mini-MDAX® (FSMX)",
     23,
     'The Excel file has the Product Title "Mini-Futures on the MDAX", while the PSS has the Product Title "Futures on Mini-MDAX". Which is the correct one?',
     "same terms, but we will unify this and apply PSS's version",
     "Product title unified to the PSS wording per ANP response; ticker (FSMX) retained per workbook convention.",
     ""),
    ("C23",
     "Futures on ATX and ATX five Indices (FATX, FATF)",
     "Futures on ATX Index (FATX)",
     24,
     "The Excel file (Column C) mentions FATX and FATF, while the pdf file only mention FATX. Has FATF been discontinued?",
     "FATF does not exist anymore, will delete it",
     "FATF (ATX five) discontinued per ANP response; product cell reduced to FATX only.",
     ""),
    ("C32",
     "Futures on FTSE Bitcoin Index",
     "Futures and Options on FTSE Crypto Indexes",
     33,
     'In the PSS, the title is listed as "Futures and Options on FTSE Crypto Indexes" while here, it says FTSE Bitcoin Index. Hence, it should potentially be renamed?',
     "Yes, Bitcoin should be deleted, and enhanced with Options",
     "Renamed to the PSS product title per ANP response.",
     'Related row "Options on FTSE Bitcoin Index" (target row 63) carried no audit comment and was left unchanged.'),
    ("E43",
     None,
     "PSS - Equity Index 02",
     44,
     "Is there currently a liquidity provisioning rebate related to this line?",
     "A program does exist, it should be referred to Equity Index 02 PSS",
     "Liquidity Provisioning Rebate reference added per ANP response (MSCI Options -> PSS Equity Index 02).",
     "Validity period of PSS Equity Index 02 not stated in the audit response - verify against the PSS before publication."),
    ("C75",
     "Dividend Futures on EURO STOXX Banks Index (FEBD) and STOXX Europe 600 Banks Index (FSBD)",
     "Dividend Futures on EURO STOXX Banks Index (FEBD)",
     76,
     "The Excel file (Column C) mentions both FEBD and FSBD as products while the PSS does not have FSBD. Has FSBD been discontinued?",
     "Yes, only FEBD ramained in PSS EI 43, PSS should be renamed and also the product cell here",
     "FSBD discontinued per ANP response; product cell reduced to FEBD.",
     ""),
    ("E91",
     "PSS - Fixed Income 12\nUntil 2026",
     "PSS - Fixed Income 12\nUntil 31.12.2026",
     92,
     '"Until 2026" is imprecise. ANP confirmed that the PSS is still valid.',
     "yes, will change to valid to 31.12.2026",
     "Imprecise validity 'Until 2026' replaced by exact date per ANP response.",
     ""),
    ("G91",
     "PSS - Fixed Income 12\nUntil 2026",
     "PSS - Fixed Income 12\nUntil 31.12.2026",
     92,
     '"Until 2026" is imprecise. ANP confirmed that the PSS is still valid.',
     "yes, will change to valid to 31.12.2026",
     "Imprecise validity 'Until 2026' replaced by exact date per ANP response (same finding as E91).",
     ""),
]

# (target_row, product_cell_text, audit_row, ia_comment, anp_response, extra_note, reason)
DELETES = [
    (13, "EURO STOXX® Banks Futures during Asian trading hours (FESB)", 14,
     "Is there currently a liquidity provisioning rebate related to this line? If not, any reason to keep it?",
     "no program currently, can be deleted", "",
     "No current LP program; row removed per ANP response."),
    (28, "Micro Futures on DAX® and EURO STOXX 50® during Asian trading hours", 29,
     "Is there currently a liquidity provisioning rebate related to this line? If not, any reason to keep it?",
     "no program currently, can be deleted", "",
     "No current LP program; row removed per ANP response."),
    (29, "Futures on STOXX® Europe 600 Factor Indices", 30,
     "As mentioned in the meeting, Axioma was a merger, and should potentially be renamed?",
     "yes, will update as in the PSS EI 62",
     "No longer exists https://www.eurex.com/ex-en/find/circulars/circular-4988296",
     "Product no longer exists per late audit note (Eurex circular 4988296); the later delisting note supersedes the earlier rename response."),
    (30, "Futures on STOXX® USA 500 Factor Futures", 31,
     "", "",
     "No longer exists https://www.eurex.com/ex-en/find/circulars/circular-4988296",
     "Product no longer exists per late audit note (Eurex circular 4988296)."),
    (55, "Eurex Daily Futures on KOSPI 200 Weekly Options (OKW1/3/4/5)", 56,
     "As mentioned in the meeting, this product has been discontinued and so, should potentially be removed from the file?",
     "no program currently, can be deleted", "",
     "Product discontinued; row removed per ANP response."),
    (70, "Equity & Basket Total Return Futures", 71,
     "Is there currently a liquidity provisioning rebate related to this line? If not, any reason to keep it?",
     "no program currently, can be deleted", "",
     "No current LP program; row removed per ANP response."),
    (78, "other Index Dividend Futures", 79,
     "Is there currently a liquidity provisioning rebate related to this line? If not, any reason to keep it?",
     "no program currently, can be deleted", "",
     "No current LP program; row removed per ANP response."),
    (94, "Short-Term Euro BTP Futures", 95,
     "ANP: line is outdated, to be removed",
     "yes, will remove", "",
     "Line outdated (PSS Fixed Income 23 validity ended 31.12.2025); row removed per ANP response."),
    (107, "Futures on Bloomberg Indices", 108,
     "Is there currently a liquidity provisioning rebate related to this line? If not, any reason to keep it?",
     "yes, will remove", "",
     "Row removed per ANP response."),
]

# Commented rows where the ANP response was a clarification only -> no change.
CLARIFICATIONS = [
    (8, "All Options (Equity Options)", 9,
     "Is Price List ECAG Section 3.2.1.1 b. generally applicable to equity options? ...",
     "Equity 01 PSS is refering to the Price List's §3.2.1.1",
     "Clarification only - no cell change required."),
    (65, "other Equity Index Options", 66,
     "Is Price List ECAG Section 3.2.1.1 a. only applicable to other equity index options?",
     "Related to other Equity Index options not specified in cell 55 - 65.",
     "Clarification only - no cell change required."),
    (84, "All Equity ETFs (ETF Options)", 85,
     "Is Price List ECAG Section 3.2.1.1 d applicable? Is it contained in the specified PSS?",
     "No, LP rebates are defined in the Equity PSS 04",
     "Clarification only - no cell change required."),
    (85, "Crypto Currency ETFs", 86,
     "Is Price List ECAG Section 3.2.1.1 d applicable? Is it contained in the specified PSS?",
     "No, LP rebates are defined in the Equity PSS 04",
     "Clarification only - no cell change required."),
    (100, "other Interest Rate Options", 101,
     "Does Price List ECAG Section 3.2.1.1 c. only relate to other interest rate options?",
     "Related to other Interest rate options not specified in cell 96 - 100.",
     "Clarification only - no cell change required."),
    (106, "Xetra-Gold Options", 107,
     "Is Price List ECAG Section 3.2.1.1 d applicable? Is it contained in the specified PSS?",
     "No, LP rebates are defined in mentioned PSS",
     "Clarification only - no cell change required."),
    (110, "Options on ETCs (Exchange-traded Commodities (ETC) Options)", 111,
     "Is Price List ECAG Section 3.2.1.1 d applicable? Is it contained in the specified PSS?",
     "No, LP rebates are defined in the Equity PSS 04",
     "Clarification only - no cell change required."),
]

# Typos noticed during review but NOT covered by any audit comment -> left unchanged.
OBSERVATIONS = [
    ("E9/G9", "MSCI Futures", "'Untill 31.12.2026' / '(untill 31.12.2026)' - spelling 'Untill'."),
    ("E7", "Stock Tracking Futures", "'Until Further notice' - inconsistent capitalisation vs 'Until further notice' elsewhere."),
    ("E87/G87", "Options on Futures on VSTOXX Index (OVS2)", "'Until -31.12.2026' - stray hyphen."),
    ("G64", "Micro-Options on DAX (ODXS)", "'(01.01.206 - 31.12.2026)' - year typo '206'."),
    ("I103", "Options on FX Futures", "'Stipends PSS - Forgein Exchnage 03' - spelling 'Forgein Exchnage'."),
    ("C63", "Options on FTSE Bitcoin Index", "Row kept unchanged (no audit comment), although the futures row for the same PSS (Equity Index 69) was renamed to the PSS title per audit row 33."),
]


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.worksheets[0]
    assert ws.title == "Transaction Fees and Incent 26", ws.title

    # --- 1. capture structure ---------------------------------------------
    merges = [str(r) for r in ws.merged_cells.ranges]
    heights = {r: d.height for r, d in ws.row_dimensions.items() if d.height is not None}
    old_filter = ws.auto_filter.ref

    # --- 2. verify + apply edits (original coordinates) -------------------
    # Style for the added E43 value: copy from populated sibling E44.
    for cell, expected, new, *_ in EDITS:
        cur = ws[cell].value
        if expected is None:
            if cur not in (None, "", " "):
                sys.exit(f"ABORT: {cell} expected empty, found {cur!r}")
        elif cur != expected:
            sys.exit(f"ABORT: {cell} expected {expected!r}, found {cur!r}")
    for cell, expected, new, *_ in EDITS:
        ws[cell] = new
    # copy style so the added reference matches its column block
    src_style = ws["E44"]
    dst = ws["E43"]
    dst.font = copy.copy(src_style.font)
    dst.alignment = copy.copy(src_style.alignment)
    dst.number_format = src_style.number_format
    # keep E43's own border/fill (grid already consistent)

    # --- 3. verify + delete rows bottom-up --------------------------------
    del_rows = sorted([d[0] for d in DELETES])
    for trow, ptext, *_ in DELETES:
        cur = ws.cell(row=trow, column=3).value  # column C = product
        if (cur or "").strip() != ptext.strip():
            sys.exit(f"ABORT: row {trow} product expected {ptext!r}, found {cur!r}")

    # unmerge everything first; re-apply adjusted ranges after deletion
    for m in merges:
        ws.unmerge_cells(m)
    for r in sorted(del_rows, reverse=True):
        ws.delete_rows(r, 1)

    def remap(row):
        """old row -> new row (None if deleted)."""
        if row in del_rows:
            return None
        return row - sum(1 for d in del_rows if d < row)

    # --- 4. re-apply merges, shrunk/shifted -------------------------------
    for m in merges:
        min_col, min_row, max_col, max_row = range_boundaries(m)
        # shrink: drop deleted rows inside the range
        rows_alive = [r for r in range(min_row, max_row + 1) if r not in del_rows]
        if not rows_alive:
            continue
        new_min, new_max = remap(rows_alive[0]), remap(rows_alive[-1])
        if new_min == new_max and min_col == max_col:
            continue  # collapsed to a single cell
        ws.merge_cells(start_row=new_min, start_column=min_col,
                       end_row=new_max, end_column=max_col)

    # --- 5. re-apply row heights ------------------------------------------
    for r in list(ws.row_dimensions.keys()):
        del ws.row_dimensions[r]
    for old_r, h in heights.items():
        new_r = remap(old_r)
        if new_r is not None:
            ws.row_dimensions[new_r].height = h

    # --- 5b. re-sync hyperlink refs (delete_rows moves cells but leaves
    # each cell.hyperlink.ref pointing at the old coordinate) ---------------
    for row in ws.iter_rows():
        for c in row:
            if getattr(c, "hyperlink", None) is not None:
                c.hyperlink.ref = c.coordinate

    # --- 6. filter range ---------------------------------------------------
    n_del = len(del_rows)
    fc, fr, lc, lr = range_boundaries(old_filter)
    ws.auto_filter.ref = f"{get_column_letter(fc)}{fr}:{get_column_letter(lc)}{lr - n_del}"

    wb.save(OUT_WB)
    print("saved:", OUT_WB)

    # --- 7. audit report ---------------------------------------------------
    rep = openpyxl.Workbook()
    s = rep.active
    s.title = "Change Log"
    hdr = ["#", "Classification", "Sheet", "Target cell/row (original)", "Row after correction",
           "Product", "Old value", "New value", "IA comment (audit col J)",
           "ANP response (audit col K)", "Late audit note (audit col N)",
           "Audit file row", "Reason", "Review note", "Confidence"]
    s.append(hdr)
    bold = Font(bold=True)
    for c in s[1]:
        c.font = bold
    i = 0
    for cell, expected, new, arow, j, k, reason, note in EDITS:
        i += 1
        col = cell[0]
        row = int(cell[1:])
        cls = "ADDED" if expected is None else ("CORRECTED" if "31.12.2026" in new and expected and "2026" in expected else "UPDATED")
        s.append([i, cls, "Transaction Fees and Incent 26", cell, remap(row),
                  ws.cell(row=remap(row), column=3).value if col != "C" else new,
                  expected or "(empty)", new, j, k, "", arow, reason, note,
                  "HIGH" if not note else "MEDIUM"])
    for trow, ptext, arow, j, k, n, reason in DELETES:
        i += 1
        s.append([i, "REMOVED", "Transaction Fees and Incent 26", f"row {trow}", "(deleted)",
                  ptext, "(entire row)", "", j, k, n, arow, reason, "", "HIGH"])
    for trow, ptext, arow, j, k, reason in CLARIFICATIONS:
        i += 1
        s.append([i, "NO_CHANGE_CLARIFIED", "Transaction Fees and Incent 26", f"row {trow}",
                  remap(trow), ptext, "(unchanged)", "", j, k, "", arow, reason, "", "HIGH"])
    for loc, prod, obs in OBSERVATIONS:
        i += 1
        s.append([i, "OBSERVATION_NOT_CHANGED", "Transaction Fees and Incent 26", loc, "",
                  prod, "", "", "", "", "", "", obs,
                  "No audit comment on this row - left unchanged per instruction that uncommented rows must not be modified.", ""])
    for col_cells in s.columns:
        letter = col_cells[0].column_letter
        s.column_dimensions[letter].width = min(60, max(10, max(len(str(c.value or "")) for c in col_cells) + 2))
    for row_cells in s.iter_rows(min_row=2):
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    m = rep.create_sheet("Summary")
    meta = [
        ("Execution timestamp", datetime.datetime.now().isoformat(timespec="seconds")),
        ("Source workbook", SRC),
        ("Output workbook", OUT_WB),
        ("Audit findings source", "overview_transaction_fees_incentives_en_answered_Audit_Findings.xlsx (col J = IA comments, col K = ANP responses, col N = late notes; audit row N = source row N-1)"),
        ("Worksheet corrected", "Transaction Fees and Incent 26 (first sheet only; Collateral / OTC IRD / Repo untouched)"),
        ("Cells edited", len(EDITS)),
        ("Rows removed", len(DELETES)),
        ("Commented rows with no change (clarifications)", len(CLARIFICATIONS)),
        ("Observations (typos etc., not changed - no audit comment)", len(OBSERVATIONS)),
        ("Rows before / after", f"110 / {110 - len(DELETES)}"),
        ("Review items", "1 - validity period for added 'PSS - Equity Index 02' (MSCI Options) to be verified against the PSS."),
        ("Assumption", "For 'Futures on STOXX Europe 600 Factor Indices' the late audit note 'No longer exists' (circular 4988296) supersedes the earlier ANP response to rename; both Factor rows were removed."),
    ]
    for k, v in meta:
        m.append([k, v])
    for c in m["A"]:
        c.font = bold
    m.column_dimensions["A"].width = 45
    m.column_dimensions["B"].width = 110
    for row_cells in m.iter_rows():
        for c in row_cells:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    rep.save(OUT_REPORT)
    print("saved:", OUT_REPORT)


if __name__ == "__main__":
    main()
